from openpyxl import load_workbook
import datetime

workbook = load_workbook('/Users/tajihiro/Works/Python/Python/更新対象.xlsx')
print (workbook.get_sheet_names())

worksheet = workbook["Sheet3"]
now = datetime.datetime.now()

for row in range(1,12):
  for col in range(1,5):
    if worksheet.cell(row=row, column=2).value == "BF00007":
      print(worksheet.cell(row = row, column = col).value)
      worksheet.cell(row=row, column=4).value = now.strftime("%Y年%m月%d日")

workbook.save('/Users/tajihiro/Works/Python/Python/更新対象.xlsx')